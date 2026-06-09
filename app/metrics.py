from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import tempfile
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DETAIL_SHEETS = [
    "抖音极速拍档",
    "抖音驭见星豪华",
    "视频号星际领航员",
    "抖音星空拍车show",
    "抖音全民驾值观",
    "视频号星空拍车show",
    "视频号全民驾值观",
    "小红书",
]

EXCLUDED_SHEETS = [
    "总月度数据表",
    "涨粉计划",
    "月度爆款汇总表",
    "各账号月度数据明细",
    "月度爆款登记",
    "部门协同",
    "3月短引直台账",
    "看后搜情况",
]

TITLE_FIELDS = ["视频名称", "视频描述"]
PUBLISH_FIELDS = ["发布时间"]
EXPOSURE_FIELDS = ["播放量"]
COMPLETION_5S_FIELDS = ["5s完播率", "5S完播率", "5s 完播率", "5S 完播率"]
INVALID_ACTORS = {"", "无", "暂无", "null", "none", "None", "NULL", "-"}
ACTOR_SPLIT_RE = re.compile(r"[、，,/\\|\s]+")
DEFAULT_READ_RANGE = "A1:AZ2000"


def run_univer_command(args: list[str], timeout: int = 180, retries: int = 0) -> subprocess.CompletedProcess[str]:
    result: subprocess.CompletedProcess[str] | None = None
    for attempt in range(retries + 1):
        result = subprocess.run(args, text=True, capture_output=True, timeout=timeout)
        if result.returncode == 0:
            return result
        combined = f"{result.stderr}\n{result.stdout}".lower()
        retryable = "timeout" in combined or "session.loadworkbookdatareplaypack" in combined
        if not retryable or attempt == retries:
            return result
        time.sleep(2 * (attempt + 1))
    assert result is not None
    return result


def compact_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def header_index(headers: list[str], candidates: list[str]) -> int | None:
    normalized = {compact_header(header): index for index, header in enumerate(headers)}
    for candidate in candidates:
        key = compact_header(candidate)
        if key in normalized:
            return normalized[key]
    return None


def actor_index(headers: list[str]) -> int | None:
    for index, header in enumerate(headers):
        if "视频演员" in str(header or ""):
            return index
    return None


def platform_from_sheet(sheet: str) -> str:
    if "抖音" in sheet:
        return "抖音"
    if "视频号" in sheet:
        return "视频号"
    if "小红书" in sheet:
        return "小红书"
    return "未识别"


def clean_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def value_at(values: list[Any], index: int | None) -> Any:
    if index is None or index >= len(values):
        return None
    return values[index]


def parse_number(value: Any) -> int:
    text = clean_text(value)
    if not text:
        return 0
    cleaned = text.replace(",", "").replace("，", "").replace(" ", "")
    try:
        return int(round(float(cleaned)))
    except ValueError:
        return 0


def parse_rate(value: Any) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    is_percent = text.endswith("%")
    cleaned = text.replace("%", "").replace(",", "").replace("，", "").strip()
    try:
        rate = float(cleaned)
    except ValueError:
        return None
    if is_percent:
        return rate / 100
    if 0 <= rate <= 1:
        return rate
    if 1 < rate <= 100:
        return rate / 100
    return None


def split_actors(value: Any) -> list[str]:
    text = clean_text(value)
    if text in INVALID_ACTORS:
        return []
    actors = [part.strip() for part in ACTOR_SPLIT_RE.split(text) if part.strip()]
    return [actor for actor in actors if actor not in INVALID_ACTORS]


def is_effective_video(title: str, exposure_value: Any, publish_time: Any) -> bool:
    return bool(clean_text(title) or clean_text(exposure_value) or clean_text(publish_time))


def normalize_records(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    records: list[dict[str, Any]] = []
    quality: dict[str, Any] = {
        "used_sheets": [],
        "excluded_sheets": EXCLUDED_SHEETS,
        "missing_fields": defaultdict(list),
        "sheet_issues": defaultdict(list),
        "invalid_values": {},
        "sheets_without_5s_completion": [],
        "notes": ["视频号 5S 未提供", "演员曝光不平摊"],
    }

    for row in rows:
        sheet = row["source_sheet"]
        headers = [clean_text(header) for header in row["headers"]]
        values = row["values"]
        title_idx = header_index(headers, TITLE_FIELDS)
        publish_idx = header_index(headers, PUBLISH_FIELDS)
        exposure_idx = header_index(headers, EXPOSURE_FIELDS)
        completion_idx = header_index(headers, COMPLETION_5S_FIELDS)
        actor_idx = actor_index(headers)

        if completion_idx is None and "5S完播率" not in quality["missing_fields"][sheet]:
            quality["missing_fields"][sheet].append("5S完播率")
            quality["sheets_without_5s_completion"].append(sheet)

        title = clean_text(value_at(values, title_idx))
        exposure_raw = value_at(values, exposure_idx)
        publish_time = clean_text(value_at(values, publish_idx))
        if not is_effective_video(title, exposure_raw, publish_time):
            continue

        actors = split_actors(value_at(values, actor_idx))
        records.append(
            {
                "platform": platform_from_sheet(sheet),
                "account_name": sheet,
                "video_title": title,
                "publish_time": publish_time or "未知日期",
                "exposure": parse_number(exposure_raw),
                "completion_rate": parse_rate(value_at(values, completion_idx)) if completion_idx is not None else None,
                "completion_metric_type": "5S完播率" if completion_idx is not None else "未提供",
                "actors": actors,
            }
        )
        if sheet not in quality["used_sheets"]:
            quality["used_sheets"].append(sheet)

    for sheet, issues in list(quality["sheet_issues"].items()):
        if not issues:
            del quality["sheet_issues"][sheet]
    quality["missing_fields"] = dict(quality["missing_fields"])
    quality["sheet_issues"] = dict(quality["sheet_issues"])
    quality["sheets_without_5s_completion"] = sorted(set(quality["sheets_without_5s_completion"]))
    return records, quality


def weighted_rate(records: list[dict[str, Any]]) -> float | None:
    weighted_sum = 0.0
    weighted_exposure = 0
    plain_rates: list[float] = []
    for record in records:
        rate = record.get("completion_rate")
        if rate is None:
            continue
        exposure = int(record.get("exposure") or 0)
        plain_rates.append(rate)
        if exposure > 0:
            weighted_sum += exposure * rate
            weighted_exposure += exposure
    if weighted_exposure:
        return weighted_sum / weighted_exposure
    if plain_rates:
        return sum(plain_rates) / len(plain_rates)
    return None


def rate_display(rate: float | None) -> str:
    if rate is None:
        return "未提供"
    return f"{rate * 100:.1f}%"


def build_account_metrics(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[record["account_name"]].append(record)
    accounts: list[dict[str, Any]] = []
    for account_name, account_records in grouped.items():
        exposure = sum(int(record["exposure"]) for record in account_records)
        rate = weighted_rate(account_records)
        completion_sources = sorted(set(record["completion_metric_type"] for record in account_records))
        accounts.append(
            {
                "platform": account_records[0]["platform"],
                "account_name": account_name,
                "video_count": len(account_records),
                "exposure": exposure,
                "average_exposure": round(exposure / len(account_records), 2) if account_records else 0,
                "completion_5s_display": rate_display(rate),
                "completion_field_source": " / ".join(completion_sources),
                "actor_video_count": sum(1 for record in account_records if record["actors"]),
                "missing_actor_video_count": sum(1 for record in account_records if not record["actors"]),
            }
        )
    return sort_account_metrics(accounts)


def sort_account_metrics(accounts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(accounts, key=lambda item: (-item["exposure"], item["account_name"]))


def ensure_detail_sheet_accounts(account_metrics: list[dict[str, Any]], quality: dict[str, Any]) -> list[dict[str, Any]]:
    accounts = list(account_metrics)
    present_accounts = {item["account_name"] for item in accounts}
    missing_fields = quality.setdefault("missing_fields", {})
    sheet_issues = quality.setdefault("sheet_issues", {})
    sheets_without_5s = set(quality.setdefault("sheets_without_5s_completion", []))
    used_sheets = quality.setdefault("used_sheets", [])

    for sheet in DETAIL_SHEETS:
        if sheet in present_accounts:
            continue
        if sheet not in used_sheets:
            used_sheets.append(sheet)
        issues = sheet_issues.setdefault(sheet, [])
        if "无有效视频行" not in issues:
            issues.append("无有效视频行")
        if "视频号" in sheet:
            fields = missing_fields.setdefault(sheet, [])
            if "5S完播率" not in fields:
                fields.append("5S完播率")
            sheets_without_5s.add(sheet)
        accounts.append(
            {
                "platform": platform_from_sheet(sheet),
                "account_name": sheet,
                "video_count": 0,
                "exposure": 0,
                "average_exposure": 0,
                "completion_5s_display": "未提供",
                "completion_field_source": "无有效视频行",
                "actor_video_count": 0,
                "missing_actor_video_count": 0,
            }
        )

    quality["sheets_without_5s_completion"] = sorted(sheets_without_5s)
    return sort_account_metrics(accounts)


def build_actor_metrics(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for record in records:
        for actor in record.get("actors", []):
            bucket = grouped.setdefault(
                actor,
                {"actor_name": actor, "video_count": 0, "accounts": set(), "contributed_exposure": 0},
            )
            bucket["video_count"] += 1
            bucket["accounts"].add(record["account_name"])
            bucket["contributed_exposure"] += int(record["exposure"])
    actors: list[dict[str, Any]] = []
    for bucket in grouped.values():
        accounts = sorted(bucket["accounts"])
        actors.append(
            {
                "actor_name": bucket["actor_name"],
                "video_count": bucket["video_count"],
                "account_count": len(accounts),
                "contributed_exposure": bucket["contributed_exposure"],
                "accounts": "、".join(accounts),
            }
        )
    return sorted(actors, key=lambda item: (-item["video_count"], -item["contributed_exposure"], item["actor_name"]))


def video_ranking_item(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "platform": record["platform"],
        "account_name": record["account_name"],
        "video_title": re.sub(r"\s+", " ", clean_text(record["video_title"])),
        "publish_time": record["publish_time"],
        "exposure": int(record["exposure"]),
        "actors": "、".join(record.get("actors") or []),
    }


def build_video_rankings(records: list[dict[str, Any]]) -> dict[str, dict[str, Any] | None]:
    titled_records = [record for record in records if clean_text(record.get("video_title"))]
    if not titled_records:
        return {"top": None, "bottom": None}
    top = max(titled_records, key=lambda record: int(record.get("exposure") or 0))
    bottom = min(titled_records, key=lambda record: int(record.get("exposure") or 0))
    return {"top": video_ranking_item(top), "bottom": video_ranking_item(bottom)}


def build_overview(records: list[dict[str, Any]], actor_metrics: list[dict[str, Any]]) -> dict[str, Any]:
    rate = weighted_rate(records)
    return {
        "total_video_count": len(records),
        "total_exposure": sum(int(record["exposure"]) for record in records),
        "overall_5s_completion_rate": rate,
        "overall_5s_completion_rate_display": rate_display(rate),
        "actor_video_count": sum(1 for record in records if record.get("actors")),
        "actor_count": len(actor_metrics),
    }


def rows_from_openpyxl(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        sheet_name = sheet.title
        if sheet_name in EXCLUDED_SHEETS:
            continue
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        sheet_rows = sheet.iter_rows(values_only=True)
        headers = next(sheet_rows, None)
        if not headers:
            continue
        for values in sheet_rows:
            rows.append({"source_sheet": sheet_name, "headers": list(headers), "values": list(values)})
    return rows


def rows_from_univer(workbook_path: Path) -> list[dict[str, Any]]:
    if shutil.which("univer") is None:
        return []
    with tempfile.TemporaryDirectory() as tmpdir:
        univer_path = Path(tmpdir) / "source.univer"
        imported = run_univer_command(["univer", "import", str(workbook_path), str(univer_path), "--json"], timeout=300, retries=1)
        if imported.returncode != 0:
            return []
        rows: list[dict[str, Any]] = []
        for sheet in DETAIL_SHEETS:
            inspected = run_univer_command(["univer", "inspect", "sheet", str(univer_path), "--sheet", sheet], timeout=180, retries=2)
            used_range = DEFAULT_READ_RANGE
            match = re.search(r"Used Range:\s*([A-Z]+[0-9]+:[A-Z]+[0-9]+)", inspected.stdout)
            if match:
                used_range = match.group(1)
            piped = run_univer_command(
                ["univer", "pipe", "out", str(univer_path), "--range", f"{sheet}!{used_range}", "--format", "json"],
                timeout=180,
                retries=2,
            )
            if piped.returncode != 0:
                continue
            matrix = json.loads(piped.stdout)
            if not matrix:
                continue
            headers = matrix[0]
            for values in matrix[1:]:
                rows.append({"source_sheet": sheet, "headers": headers, "values": values})
        return rows


def build_dataset_from_workbook(workbook_path: Path, source_file_name: str) -> dict[str, Any]:
    rows = rows_from_openpyxl(workbook_path)
    if not rows and os.environ.get("XINGTU_ENABLE_UNIVER_FALLBACK") == "1":
        rows = rows_from_univer(workbook_path)
    records, quality = normalize_records(rows)
    actor_metrics = build_actor_metrics(records)
    account_metrics = ensure_detail_sheet_accounts(build_account_metrics(records), quality)
    overview = build_overview(records, actor_metrics)
    overview.update(
        {
            "quality_status": quality_status(quality),
            "source_file_name": source_file_name,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        }
    )
    return {
        "overview": overview,
        "account_metrics": account_metrics,
        "actor_metrics": actor_metrics,
        "video_rankings": build_video_rankings(records),
        "quality_report": quality,
    }


def quality_status(quality: dict[str, Any]) -> str:
    sheets_without_5s = quality.get("sheets_without_5s_completion", [])
    if sheets_without_5s:
        return f"{len(sheets_without_5s)} 个 sheet 缺少 5S"
    missing_count = sum(len(fields) for fields in quality.get("missing_fields", {}).values())
    issue_count = sum(len(issues) for issues in quality.get("sheet_issues", {}).values())
    total = missing_count + issue_count
    if total == 0:
        return "数据正常"
    return f"{total} 个质量提示"
