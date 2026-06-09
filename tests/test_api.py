import unittest
from tempfile import NamedTemporaryFile

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


class ApiTest(unittest.TestCase):
    def test_overview_contract_contains_confirmed_kpis(self):
        client = TestClient(app)

        response = client.get("/api/short-video/overview")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("overview", payload)
        self.assertIn("account_metrics", payload)
        self.assertIn("actor_metrics", payload)
        self.assertIn("video_rankings", payload)
        self.assertIn("quality_report", payload)
        self.assertEqual(payload["module"]["slug"], "xingtu")
        self.assertEqual(payload["overview"]["total_video_count"], 320)
        self.assertEqual(payload["overview"]["total_exposure"], 12226000)
        self.assertIn("video_title", payload["video_rankings"]["top"])
        self.assertIn("video_title", payload["video_rankings"]["bottom"])

    def test_modules_endpoint_exposes_hub_entries(self):
        client = TestClient(app)

        response = client.get("/api/modules")

        self.assertEqual(response.status_code, 200)
        modules = {item["slug"]: item for item in response.json()["modules"]}
        self.assertEqual(modules["xingtu"]["dashboard_path"], "/xingtu")
        self.assertEqual(modules["xingtu"]["admin_path"], "/admin/xingtu")
        self.assertTrue(modules["xingtu"]["upload_enabled"])
        self.assertEqual(modules["oae"]["dashboard_path"], "/oae")
        self.assertFalse(modules["oae"]["upload_enabled"])

    def test_module_overview_loads_oae_dashboard_source_metrics(self):
        client = TestClient(app)

        response = client.get("/api/bi/oae/overview")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["module"]["slug"], "oae")
        self.assertEqual(payload["overview"]["module_status"], "ready")
        self.assertEqual(payload["overview"]["source_contract"], "feishu_dashboard_source_tsv")
        self.assertEqual(payload["overview"]["report_date"], "2026-06-08")
        self.assertGreater(payload["overview"]["total_exposure"], 0)
        self.assertIn("oae_dashboard", payload)
        self.assertGreater(len(payload["oae_dashboard"]["lead_accounts"]), 0)

    def test_oae_route_serves_original_operations_daily_bi_shell(self):
        client = TestClient(app)

        response = client.get("/oae")

        self.assertEqual(response.status_code, 200)
        html = response.text
        self.assertIn("运营日报 BI", html)
        self.assertIn("今日判断", html)
        self.assertIn('data-dashboard-mode="business"', html)
        self.assertIn("/dashboard/daily/latest", html)

    def test_oae_original_dashboard_api_contract(self):
        client = TestClient(app)

        response = client.get("/dashboard/daily/latest")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["report_date"], "2026-06-08")
        self.assertEqual(payload["source"]["type"], "feishu_dashboard_source_tsv")
        self.assertIn("raw_leads", payload["overview"])
        self.assertGreater(payload["overview"]["impressions"]["actual"], 0)
        self.assertIn("lead_anchors", payload)
        self.assertIn("seed_anchors", payload)

    def test_oae_original_dashboard_trends_api_contract(self):
        client = TestClient(app)

        response = client.get("/dashboard/daily/trends?end_date=2026-06-08")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["date_range"]["end"], "2026-06-08")
        self.assertTrue(payload["daily_trends"])
        self.assertTrue(payload["core_kpi_summary"])
        self.assertIn("monthly_comparison", payload)

    def test_oae_upload_is_blocked_until_source_contract_exists(self):
        client = TestClient(app)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "OAE测试"
        sheet.append(["字段", "数值"])
        sheet.append(["曝光", 1000])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            response = client.post(
                "/api/bi/oae/admin/upload",
                files={"file": ("oae.xlsx", tmp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        self.assertEqual(response.status_code, 422)
        self.assertIn("只读数据源", response.json()["detail"])

    def test_unknown_module_returns_404(self):
        client = TestClient(app)

        response = client.get("/api/bi/unknown/overview")

        self.assertEqual(response.status_code, 404)

    def test_upload_endpoint_accepts_excel_file(self):
        client = TestClient(app)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "抖音测试账号"
        sheet.append(["视频名称", "发布时间", "播放量", "5s完播率", "视频演员"])
        sheet.append(["测试视频", "2026/3/1", 1000, 0.2, "桂婕、曹嘉洋"])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            response = client.post(
                "/api/admin/upload",
                files={"file": ("sample_short_video.xlsx", tmp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "ready")
        self.assertGreaterEqual(payload["overview"]["total_video_count"], 1)

    def test_upload_endpoint_rejects_empty_dataset_without_overwriting_current_data(self):
        client = TestClient(app)
        before = client.get("/api/short-video/overview").json()["overview"]["total_video_count"]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "抖音测试账号"
        sheet.append(["视频名称", "发布时间", "播放量", "5s完播率", "视频演员"])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            response = client.post(
                "/api/admin/upload",
                files={"file": ("empty_short_video.xlsx", tmp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        self.assertEqual(response.status_code, 422)
        self.assertIn("未识别到有效视频行", response.json()["detail"])
        after = client.get("/api/short-video/overview").json()["overview"]["total_video_count"]
        self.assertEqual(after, before)


if __name__ == "__main__":
    unittest.main()
